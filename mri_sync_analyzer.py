#!/usr/bin/env python3
"""
MRI Audio-Video Synchronization Analyzer
Analyzes 2,371 MRI videos for audio-video sync issues, missing audio, duration mismatches.
Generates detailed Excel reports and visualizations.
"""

import os
import subprocess
import json
from pathlib import Path
import cv2
import librosa
import numpy as np
from scipy import signal
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')


class MRISyncAnalyzer:
    """Analyzes audio-video synchronization in MRI video dataset."""
    
    def __init__(self, dataset_path, output_dir="Reports"):
        """
        Initialize analyzer.
        
        Args:
            dataset_path: Path to folder containing MRI videos
            output_dir: Directory for output files
        """
        self.dataset_path = Path(dataset_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        self.results = []
        self.sync_offsets = []
        self.issue_categories = {}
        
    def get_video_metadata(self, video_path):
        """Extract video duration and fps using OpenCV."""
        try:
            cap = cv2.VideoCapture(str(video_path))
            if not cap.isOpened():
                return None, None, None
            
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            duration = frame_count / fps if fps > 0 else 0
            
            cap.release()
            return duration, fps, frame_count
        except Exception as e:
            print(f"Error reading video {video_path}: {e}")
            return None, None, None
    
    def extract_audio(self, video_path):
        """Extract audio from video using ffmpeg."""
        try:
            # Get audio info first
            probe_cmd = [
                'ffprobe', '-v', 'error',
                '-select_streams', 'a:0',
                '-show_entries', 'stream=codec_type',
                '-of', 'default=noprint_wrappers=1:nokey=1:noprint_wrappers=1',
                str(video_path)
            ]
            result = subprocess.run(probe_cmd, capture_output=True, text=True, timeout=5)
            
            if not result.stdout.strip():
                return None, None, "NO_AUDIO_STREAM"
            
            # Extract audio to temporary WAV
            audio_path = f"/tmp/temp_audio_{hash(str(video_path)) % 100000}.wav"
            extract_cmd = [
                'ffmpeg', '-i', str(video_path),
                '-q:a', '9', '-n',
                audio_path
            ]
            subprocess.run(extract_cmd, capture_output=True, timeout=30)
            
            if not os.path.exists(audio_path):
                return None, None, "EXTRACTION_FAILED"
            
            # Load audio
            y, sr = librosa.load(audio_path, sr=None)
            duration = len(y) / sr
            
            # Check if audio is silent
            if np.max(np.abs(y)) < 0.01:
                return y, sr, "SILENT_AUDIO"
            
            # Clean up
            os.remove(audio_path)
            return y, sr, "SUCCESS"
            
        except subprocess.TimeoutExpired:
            return None, None, "TIMEOUT"
        except Exception as e:
            return None, None, f"ERROR: {str(e)[:30]}"
    
    def calculate_video_motion(self, video_path, sample_frames=30):
        """Calculate motion energy from video frames."""
        try:
            cap = cv2.VideoCapture(str(video_path))
            if not cap.isOpened():
                return None
            
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            if frame_count == 0:
                return None
            
            # Sample frames throughout video
            frame_indices = np.linspace(0, frame_count - 1, sample_frames, dtype=int)
            motion_energy = []
            prev_frame = None
            
            for idx in frame_indices:
                cap.set(cv2.CAP_PROP_POS_FRAMES, idx)
                ret, frame = cap.read()
                
                if not ret or frame is None:
                    continue
                
                # Convert to grayscale
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                
                if prev_frame is not None:
                    # Calculate frame difference (motion)
                    diff = cv2.absdiff(gray, prev_frame)
                    motion = np.sum(diff)
                    motion_energy.append(motion)
                
                prev_frame = gray
            
            cap.release()
            
            if len(motion_energy) == 0:
                return None
            
            # Normalize
            motion_array = np.array(motion_energy)
            motion_array = (motion_array - np.min(motion_array)) / (np.max(motion_array) - np.min(motion_array) + 1e-6)
            
            return motion_array
            
        except Exception as e:
            print(f"Error calculating video motion: {e}")
            return None
    
    def calculate_sync_offset(self, audio, sr, video_motion):
        """Calculate audio-video sync offset using cross-correlation."""
        try:
            if audio is None or len(audio) == 0 or video_motion is None:
                return None
            
            # Extract audio energy envelope
            S = librosa.feature.melspectrogram(y=audio, sr=sr, n_mels=13)
            audio_energy = np.sqrt(np.sum(S**2, axis=0))
            
            # Normalize
            audio_energy = (audio_energy - np.mean(audio_energy)) / (np.std(audio_energy) + 1e-6)
            
            # Resample to match video motion sampling
            if len(audio_energy) != len(video_motion):
                audio_energy = np.interp(
                    np.linspace(0, len(audio_energy) - 1, len(video_motion)),
                    np.arange(len(audio_energy)),
                    audio_energy
                )
            
            # Cross-correlation
            correlation = signal.correlate(audio_energy, video_motion, mode='same')
            lag = signal.correlation_lags(len(audio_energy), len(video_motion), mode='same')
            
            max_corr_idx = np.argmax(np.abs(correlation))
            offset_frames = lag[max_corr_idx]
            
            # Convert to milliseconds (assuming 30fps for motion sampling)
            offset_ms = (offset_frames / 30.0) * 1000
            
            return offset_ms
            
        except Exception as e:
            print(f"Error calculating sync offset: {e}")
            return None
    
    def categorize_issue(self, video_duration, audio_duration, sync_offset, audio_status):
        """Categorize video by issue type."""
        issues = []
        
        if audio_status != "SUCCESS":
            if audio_status == "NO_AUDIO_STREAM":
                issues.append("MISSING_AUDIO")
            elif audio_status == "SILENT_AUDIO":
                issues.append("SILENT_AUDIO")
            elif audio_status == "EXTRACTION_FAILED":
                issues.append("EXTRACTION_FAILED")
            return issues
        
        # Duration mismatch
        if video_duration and audio_duration:
            duration_diff = abs(video_duration - audio_duration)
            if duration_diff > 0.1:  # More than 100ms difference
                issues.append("DURATION_MISMATCH")
        
        # Sync offset
        if sync_offset is not None:
            if sync_offset < 200:
                issues.append("GOOD_SYNC")
            elif 200 <= sync_offset < 500:
                issues.append("MINOR_DELAY")
            elif 500 <= sync_offset < 1000:
                issues.append("MAJOR_DELAY")
            else:
                issues.append("SEVERE_DELAY")
        
        return issues if issues else ["UNKNOWN"]
    
    def analyze_dataset(self):
        """Analyze all videos in dataset."""
        video_files = list(self.dataset_path.glob("*"))
        video_files = [f for f in video_files if f.suffix.lower() in ['.mp4', '.mov', '.avi', '.mkv', '.flv']]
        
        total = len(video_files)
        print(f"Found {total} video files to analyze...")
        
        for idx, video_file in enumerate(video_files, 1):
            print(f"[{idx}/{total}] Analyzing: {video_file.name}")
            
            # Get video metadata
            video_duration, fps, frame_count = self.get_video_metadata(video_file)
            
            # Extract audio
            audio, sr, audio_status = self.extract_audio(video_file)
            audio_duration = len(audio) / sr if audio is not None and sr else None
            
            # Calculate sync offset
            sync_offset = None
            if audio_status == "SUCCESS":
                video_motion = self.calculate_video_motion(video_file)
                sync_offset = self.calculate_sync_offset(audio, sr, video_motion)
            
            # Categorize issues
            issues = self.categorize_issue(video_duration, audio_duration, sync_offset, audio_status)
            primary_issue = issues[0] if issues else "UNKNOWN"
            
            # Store result
            result = {
                'filename': video_file.name,
                'filepath': str(video_file),
                'video_duration_s': round(video_duration, 2) if video_duration else None,
                'audio_duration_s': round(audio_duration, 2) if audio_duration else None,
                'sync_offset_ms': round(sync_offset, 2) if sync_offset else None,
                'fps': round(fps, 2) if fps else None,
                'frame_count': frame_count,
                'audio_status': audio_status,
                'issues': ', '.join(issues),
                'primary_issue': primary_issue
            }
            
            self.results.append(result)
            
            if sync_offset is not None:
                self.sync_offsets.append(sync_offset)
            
            # Track issue categories
            for issue in issues:
                self.issue_categories[issue] = self.issue_categories.get(issue, 0) + 1
        
        print(f"\n✅ Analysis complete! {len(self.results)} files processed.")
    
    def generate_excel_report(self, filename="MRI_Issues_Organized.xlsx"):
        """Generate detailed Excel report."""
        output_path = self.output_dir / filename
        
        df_all = pd.DataFrame(self.results)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Files organized by issue
        ws1 = wb.create_sheet("FILES_BY_ISSUE")
        ws1.append(["ISSUE TYPE", "COUNT", "FILES"])
        
        for issue, count in sorted(self.issue_categories.items(), key=lambda x: x[1], reverse=True):
            issue_files = df_all[df_all['primary_issue'] == issue]
            files_str = "; ".join(issue_files['filename'].tolist()[:5])
            if len(issue_files) > 5:
                files_str += f"; ... +{len(issue_files) - 5} more"
            ws1.append([issue, count, files_str])
        
        # Sheet 2: All files with metadata
        ws2 = wb.create_sheet("ALL_FILES")
        for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
        
        # Format headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for sheet in [ws1, ws2]:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Adjust column widths
        for sheet in [ws1, ws2]:
            for column in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        
        wb.save(str(output_path))
        print(f"✅ Excel report saved: {output_path}")
        return output_path
    
    def generate_visualizations(self, filename="MRI_Analysis_Charts.png"):
        """Generate visualization dashboard."""
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle('MRI Audio-Video Sync Analysis Dashboard', fontsize=18, fontweight='bold')
        
        # Plot 1: Issue distribution pie chart
        issue_counts = dict(sorted(self.issue_categories.items(), key=lambda x: x[1], reverse=True))
        axes[0, 0].pie(issue_counts.values(), labels=issue_counts.keys(), autopct='%1.1f%%',
                       startangle=90, colors=sns.color_palette("husl", len(issue_counts)))
        axes[0, 0].set_title('Issue Type Distribution', fontweight='bold', fontsize=12)
        
        # Plot 2: Sync offset histogram
        if self.sync_offsets:
            axes[0, 1].hist(self.sync_offsets, bins=30, color='steelblue', edgecolor='black', alpha=0.7)
            axes[0, 1].axvline(np.mean(self.sync_offsets), color='red', linestyle='--', linewidth=2, label=f'Mean: {np.mean(self.sync_offsets):.1f}ms')
            axes[0, 1].set_xlabel('Sync Offset (ms)', fontweight='bold')
            axes[0, 1].set_ylabel('Frequency', fontweight='bold')
            axes[0, 1].set_title('Sync Offset Distribution', fontweight='bold', fontsize=12)
            axes[0, 1].legend()
            axes[0, 1].grid(axis='y', alpha=0.3)
        
        # Plot 3: Issue counts bar chart
        df_all = pd.DataFrame(self.results)
        primary_issues = df_all['primary_issue'].value_counts().sort_values(ascending=False)
        axes[1, 0].barh(range(len(primary_issues)), primary_issues.values, color='coral', edgecolor='black')
        axes[1, 0].set_yticks(range(len(primary_issues)))
        axes[1, 0].set_yticklabels(primary_issues.index)
        axes[1, 0].set_xlabel('Number of Files', fontweight='bold')
        axes[1, 0].set_title('Issue Counts by Type', fontweight='bold', fontsize=12)
        axes[1, 0].grid(axis='x', alpha=0.3)
        
        # Plot 4: Summary statistics
        axes[1, 1].axis('off')
        summary_text = f"""
        ANALYSIS SUMMARY
        {'='*50}
        
        Total Files Analyzed: {len(self.results)}
        
        Issue Breakdown:
        """
        for issue, count in sorted(self.issue_categories.items(), key=lambda x: x[1], reverse=True):
            pct = (count / len(self.results)) * 100
            summary_text += f"\n  • {issue}: {count} ({pct:.1f}%)"
        
        if self.sync_offsets:
            summary_text += f"\n\nSync Offset Stats:"
            summary_text += f"\n  Mean: {np.mean(self.sync_offsets):.2f} ms"
            summary_text += f"\n  Median: {np.median(self.sync_offsets):.2f} ms"
            summary_text += f"\n  Std Dev: {np.std(self.sync_offsets):.2f} ms"
            summary_text += f"\n  Range: {np.min(self.sync_offsets):.2f} - {np.max(self.sync_offsets):.2f} ms"
        
        axes[1, 1].text(0.05, 0.95, summary_text, transform=axes[1, 1].transAxes,
                       fontfamily='monospace', fontsize=10, verticalalignment='top',
                       bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        plt.tight_layout()
        output_path = self.output_dir / filename
        plt.savefig(str(output_path), dpi=300, bbox_inches='tight')
        print(f"✅ Visualizations saved: {output_path}")
        plt.close()
        
        return output_path


def main():
    """Main execution."""
    # Configuration
    DATASET_PATH = "/Users/harshkumar/Desktop/MRI Project/dataset_2drt_video_only"
    OUTPUT_DIR = "Reports"
    
    # Initialize analyzer
    analyzer = MRISyncAnalyzer(DATASET_PATH, OUTPUT_DIR)
    
    # Run analysis
    analyzer.analyze_dataset()
    
    # Generate outputs
    analyzer.generate_excel_report()
    analyzer.generate_visualizations()
    
    print("\n" + "="*60)
    print("🎉 MRI Audio-Video Sync Analysis Complete!")
    print("="*60)
    print(f"📊 Issue Summary:")
    for issue, count in sorted(analyzer.issue_categories.items(), key=lambda x: x[1], reverse=True):
        print(f"   {issue}: {count}")


if __name__ == "__main__":
    main()
